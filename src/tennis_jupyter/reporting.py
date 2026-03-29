"""Reporting utilities for CSV and Excel exports."""

from __future__ import annotations

from os import PathLike
from pathlib import Path
from typing import BinaryIO

import pandas as pd
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font

from .shared import safe_ratio


def build_excel_report_tables(grouped: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """Build export tables for the Excel workbook."""
    tables: dict[str, pd.DataFrame] = {}

    raw = grouped.copy()
    raw = raw.sort_values(
        ["Match Year", "Match Date", "player", "opp"],
        na_position="last",
    )
    tables["MatchSummary"] = raw

    year_result = (
        grouped.groupby(["Match Year", "Match Result"], dropna=False)
        .size()
        .unstack(fill_value=0)
        .reset_index()
    )
    for column in ["W", "L"]:
        if column not in year_result.columns:
            year_result[column] = 0

    year_result = year_result[["Match Year", "W", "L"]].sort_values("Match Year")
    year_result["Total Matches"] = year_result["W"] + year_result["L"]
    year_result["Win Rate"] = safe_ratio(
        year_result["W"],
        year_result["Total Matches"],
    )
    tables["Pivot_Year_Result"] = year_result

    year_games = (
        grouped.groupby("Match Year", dropna=False)[
            ["Games Won", "Games Lost", "Sets Won", "Sets Lost"]
        ]
        .sum()
        .reset_index()
        .sort_values("Match Year")
    )
    tables["Pivot_Year_Games"] = year_games

    player_overview = (
        grouped.groupby("player", dropna=False)
        .agg(
            Matches=("matchId", "nunique"),
            Wins=("Match Result", lambda values: (values == "W").sum()),
            Losses=("Match Result", lambda values: (values == "L").sum()),
            Sets_Won=("Sets Won", "sum"),
            Sets_Lost=("Sets Lost", "sum"),
            Games_Won=("Games Won", "sum"),
            Games_Lost=("Games Lost", "sum"),
        )
        .reset_index()
        .sort_values(["Matches", "Wins", "player"], ascending=[False, False, True])
    )
    player_overview["Win Rate"] = safe_ratio(
        player_overview["Wins"],
        player_overview["Matches"],
    )
    tables["Pivot_Player_Overview"] = player_overview

    player_serve_return = (
        grouped.groupby("player", dropna=False)[
            [
                "first_serve_attempt",
                "first_serve_in",
                "first_serve_won",
                "second_serve_attempt",
                "second_serve_in",
                "second_serve_won",
                "first_serve_return_opportunity",
                "first_serve_return_in",
                "first_serve_return_won",
                "second_serve_return_opportunity",
                "second_serve_return_in",
                "second_serve_return_won",
            ]
        ]
        .sum()
        .reset_index()
    )

    player_serve_return["1st Serve In %"] = safe_ratio(
        player_serve_return["first_serve_in"],
        player_serve_return["first_serve_attempt"],
    )
    player_serve_return["1st Serve Won %"] = safe_ratio(
        player_serve_return["first_serve_won"],
        player_serve_return["first_serve_in"],
    )
    player_serve_return["2nd Serve Won %"] = safe_ratio(
        player_serve_return["second_serve_won"],
        player_serve_return["second_serve_attempt"],
    )
    player_serve_return["1st Return In %"] = safe_ratio(
        player_serve_return["first_serve_return_in"],
        player_serve_return["first_serve_return_opportunity"],
    )
    player_serve_return["1st Return Won %"] = safe_ratio(
        player_serve_return["first_serve_return_won"],
        player_serve_return["first_serve_return_opportunity"],
    )
    player_serve_return["2nd Return In %"] = safe_ratio(
        player_serve_return["second_serve_return_in"],
        player_serve_return["second_serve_return_opportunity"],
    )
    player_serve_return["2nd Return Won %"] = safe_ratio(
        player_serve_return["second_serve_return_won"],
        player_serve_return["second_serve_return_opportunity"],
    )
    tables["Pivot_Serve_Return"] = player_serve_return[
        [
            "player",
            "1st Serve In %",
            "1st Serve Won %",
            "2nd Serve Won %",
            "1st Return In %",
            "1st Return Won %",
            "2nd Return In %",
            "2nd Return Won %",
        ]
    ].sort_values("player")

    return tables


def _autosize_and_format_sheet(ws) -> None:
    """Format sheet headers and percentage columns for readability."""
    percent_headers = {"Win Rate"} | {
        cell.value
        for cell in ws[1]
        if isinstance(cell.value, str) and "%" in cell.value
    }
    header_to_column = {
        cell.value: cell.column for cell in ws[1] if cell.value is not None
    }

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for column_cells in ws.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        max_length = max((len(value) for value in values), default=0)
        ws.column_dimensions[column_cells[0].column_letter].width = min(
            max(max_length + 2, 12),
            36,
        )

    for header in percent_headers:
        column_index = header_to_column.get(header)
        if not column_index:
            continue
        for row_index in range(2, ws.max_row + 1):
            ws.cell(row=row_index, column=column_index).number_format = "0.0%"


def _add_chart_title(chart, title: str) -> None:
    """Apply consistent chart sizing and title formatting."""
    chart.title = title
    chart.style = 10
    chart.width = 14
    chart.height = 8


def add_report_charts(workbook) -> None:
    """Add a summary chart sheet to the workbook."""
    if "Charts" in workbook.sheetnames:
        del workbook["Charts"]

    charts_ws = workbook.create_sheet("Charts")
    charts_ws["A1"] = "Tennis Match Summary Pivot Charts"
    charts_ws["A1"].font = Font(bold=True, size=14)

    if "Pivot_Year_Result" in workbook.sheetnames:
        ws = workbook["Pivot_Year_Result"]
        if ws.max_row >= 2 and ws.max_column >= 3:
            wl_chart = BarChart()
            _add_chart_title(wl_chart, "Matches by Year (W/L)")
            data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=ws.max_row)
            categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            wl_chart.add_data(data, titles_from_data=True)
            wl_chart.set_categories(categories)
            wl_chart.y_axis.title = "Matches"
            wl_chart.x_axis.title = "Year"
            charts_ws.add_chart(wl_chart, "A3")

            win_rate_chart = LineChart()
            _add_chart_title(win_rate_chart, "Win Rate by Year")
            rate_data = Reference(ws, min_col=5, max_col=5, min_row=1, max_row=ws.max_row)
            win_rate_chart.add_data(rate_data, titles_from_data=True)
            win_rate_chart.set_categories(categories)
            win_rate_chart.y_axis.title = "Win Rate"
            win_rate_chart.y_axis.number_format = "0%"
            win_rate_chart.x_axis.title = "Year"
            charts_ws.add_chart(win_rate_chart, "P3")

    if "Pivot_Year_Games" in workbook.sheetnames:
        ws = workbook["Pivot_Year_Games"]
        if ws.max_row >= 2 and ws.max_column >= 5:
            games_chart = LineChart()
            _add_chart_title(games_chart, "Games Won/Lost by Year")
            data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=ws.max_row)
            categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            games_chart.add_data(data, titles_from_data=True)
            games_chart.set_categories(categories)
            games_chart.y_axis.title = "Games"
            games_chart.x_axis.title = "Year"
            charts_ws.add_chart(games_chart, "A22")

    if "Pivot_Player_Overview" in workbook.sheetnames:
        ws = workbook["Pivot_Player_Overview"]
        if ws.max_row >= 2 and ws.max_column >= 4:
            chart = BarChart()
            _add_chart_title(chart, "Top 10 Players by Matches (W/L)")
            max_row = min(ws.max_row, 11)
            data = Reference(ws, min_col=2, max_col=4, min_row=1, max_row=max_row)
            categories = Reference(ws, min_col=1, min_row=2, max_row=max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            chart.y_axis.title = "Count"
            charts_ws.add_chart(chart, "P22")


def write_excel_report(grouped: pd.DataFrame, out_path: str | PathLike[str] | BinaryIO) -> None:
    """Write the summary dataset and pivot tables to Excel."""
    if isinstance(out_path, (str, PathLike)):
        out_path = Path(out_path)
        out_path.parent.mkdir(parents=True, exist_ok=True)
    tables = build_excel_report_tables(grouped)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for sheet_name, table in tables.items():
            table.to_excel(writer, sheet_name=sheet_name, index=False)

        workbook = writer.book
        for worksheet in workbook.worksheets:
            _autosize_and_format_sheet(worksheet)
        add_report_charts(workbook)
        _autosize_and_format_sheet(workbook["Charts"])
